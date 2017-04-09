from __future__ import division
from openpyxl import load_workbook
import nltk

def get_excel_data(sheet):
    headers = ["str_id", "project_id", "tweet_text", "label"]
    excel_data = []  # list of dictionaries ##NAME OF LIST of dict
    for row_num, row in enumerate(sheet):
        if row_num is 0:  # skip the first row (don't want headers)
            continue
        row_data = {}  # open a dictionary ##NAME OF DICTIONARY OF VALUES IN EACH ROW
        for col_num, cell in enumerate(
                row):  # populating the dictionary with header names as keys and cell content as values# populating the dictionary with header names as keys and cell content as values
            if col_num > len(headers) - 2:  # ask Max what this is
                continue
            key = headers[col_num]
            value = cell.value
            row_data[key] = value
        excel_data.append(row_data)  # adding each key-value pair to the excel_data list
    print(excel_data)
    return excel_data


# Returns tweet text and projid in a tuple:
def find_tweet_text(excel_data):
    tuple_tweet_projid = []
    for row_data in excel_data:
        if row_data != None:
            tweet = row_data['tweet_text'], row_data['project_id']
            tuple_tweet_projid.append(tweet)
    return tuple_tweet_projid


# find scope of negation, between [ ]:
def find_scope(tuple_tweet_projid):
    # retrieve the tweet_text only, firts element in the tuple:
    scope_output = []
    for tweet in tuple_tweet_projid:
        # separate tweet text only
        tweet_text = tweet[0]
        # ignore tweets with value None, i.e. at end of the list
        if tweet_text != None:
            # find the scope marks( [ ] ):
            if '[' and ']' in tweet_text:
                str1 = '['
                str2 = ']'
                scope_starts = tweet_text.find(str1)
                scope_ends = tweet_text.find(str2)
                # print scope_starts, scope_ends
                scope = tweet_text[scope_starts: scope_ends + 1]
                # print scopes
                proj_id = tweet[1]
                # print proj_id
                scope_with_projid = proj_id, scope
        scope_output.append(scope_with_projid)
    return scope_output

#coparing scopes:
def scope_cleaning_tokenizing(scope_output):
    #cleaning tweet_text in the first set of tuples (annotator 1):
    tweet_tokens = []
    for proj_id, tweet_text in scope_output:
        #print('tweet_text type: ', type(tweet_text))
        #eliminating scope marks ([]):
        clean_annot = tweet_text.strip('[]')
        #print('this is clean_annot: ', clean_annot)
        #tokenizing:
        tokens_annot = nltk.word_tokenize(clean_annot)
        tweet_tokens.append((proj_id, tokens_annot))
    return tweet_tokens


#iterating over both sets of tuples to find strict and partial matches between annotations:
def scope_match(scope_clean1, scope_clean2):
    for val1, val2 in zip(scope_clean1, scope_clean2):
        #1.calculating partial matches with symmetric_difference:
        partial_match = list(set(val1[1]).symmetric_difference(val2[1]))
        #2.calculating strict (or complete) overlap using intersection:
        strict_match = list(set(val1[1]).intersection(val2[1]))
        if len(partial_match) > 0:
            print('Partial match: ', val1[0], len(partial_match), partial_match)
        else:
            print('Strict match: ', val1[0], len(strict_match))
    return


if __name__ == '__main__':
    ##I'll use these raw_input functions in the final version of the program    
    # workbook_input = raw_input('enter spreadsheet name (should be xlsx): ') # open files
    # workbook = load_workbook('./' + workbook_input + '.xlsx')
    # sheet_input1 = raw_input('enter annotated spreadsheet name (as it appears on the tab name): ')
    # sheet_input2 = raw_input('enter master list spreadsheet name (as it appears on the tab name): ')
    # sheet1 = workbook[sheet_input1]
    # sheet2 = workbook[sheet_input2]


    workbook = load_workbook('./Scope-Annotation-TrainingDONE.xlsx')

    sheet1 = workbook['ScopeTrainingAnnotNORMA']
    sheet2 = workbook['ScopeTrainingAnnotBRIANA']

    excel_data1 = get_excel_data(sheet1)
    excel_data2 = get_excel_data(sheet2)

    # Extracting tweet_text and proj_id:
    tuple_tweet_projid1 = find_tweet_text(excel_data1)
    # print tuple_tweet_projid1
    tuple_tweet_projid2 = find_tweet_text(excel_data2)
    # print tuple_tweet_projid2

    # Finding scopes:
    scope_found1 = find_scope(tuple_tweet_projid1)
    # print scope_found1
    scope_found2 = find_scope(tuple_tweet_projid2)
    # print scope_found2
    scope_clean1 = scope_cleaning_tokenizing(scope_found1)
    #print scope_clean1
    scope_clean2 = scope_cleaning_tokenizing(scope_found2)
    #print scope_clean2

    # Matching scopes:
    scope_agreement = scope_match(scope_clean1, scope_clean2)
    print(scope_agreement)
