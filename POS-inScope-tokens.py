#program that extracts text from a spreadsheets, isolates a chunck of tokens from each text 
#(chunk that has been annotated in between square brackets), tokenizes and POS-tags that chunk,
#extracts the POS tag values only (to get the POS representation of the text),
#and finally sorts each chunk of text according to the number of POS values present


from __future__ import division
from openpyxl import load_workbook
import re
import nltk
from nltk import pos_tag
from operator import itemgetter

def get_excel_data(sheet):
    headers = ["str_id", "project_id", "tweet_text", "label"]
    excel_data = []  # list of dictionaries ##NAME OF LIST of dict
    for row_num, row in enumerate(sheet):
        if row_num is 0:  # skip the first row (don't want headers)
            continue
        row_data = {}  # open a dictionary ##NAME OF DICTIONARY OF VALUES IN EACH ROW
        for col_num, cell in enumerate(
                row):  # populating the dictionary with header names as keys and cell content as values# populating the dictionary with header names as keys and cell content as values
            if col_num > len(headers) - 2:  # ask Max what this is
                continue
            key = headers[col_num]
            value = cell.value
            row_data[key] = value
        excel_data.append(row_data)  # adding each key-value pair to the excel_data list
    return excel_data


# Returns tweet text only
def find_tweet_text(excel_data):
    tweet_text = []
    for row_data in excel_data:
        if row_data != None:
            tweet = row_data['tweet_text']
            #print('this is row_data[\'tweet_text\'] type: ', type(tweet))
            tweet_text.append(tweet)
    #print('this is tweet_text: ', tweet_text)
    return tweet_text


# find scope of negation, between [ ]
def find_scope(tweet_text):
    # retrieve the tweet_text only, firts element in the tuple:
    scope_output = []
    for tweet in tweet_text:
        # separate tweet text only
        #tweet_text = tweet
        # ignore tweets with value None, i.e. at end of the list
        if tweet != None:
            # find the scope marks( [ ] ):
	        if '[' and ']' in tweet:
	            str1 = '['
	            str2 = ']'
	            scope_starts = tweet.find(str1)
	            scope_ends = tweet.find(str2)
	            # print scope_starts, scope_ends
	            scope = tweet[scope_starts: scope_ends + 1]
	            # print scopes
	            #proj_id = tweet[1]
	            # print proj_id
	            #scope_with_projid = proj_id, scope
	        scope_output.append(scope)
	        #print('this is scope output: ', scope_output)
    #print(type(scope_output))
    return scope_output

#coparing scopes:
def scope_cleaning_tokenizing(scope_output):
    #cleaning tweet_text in the first set of tuples (annotator 1):
    tweet_tokens = []
    for tweet in scope_output:
        #print('tweet type: ', type(tweet))
        tweet_nob = re.sub(r'^(b\'\")', '', tweet)
        #print('tweet_nob: ', tweet_nob)
        #eliminating scope marks ([]):
        clean_annot = tweet_nob.replace('[', '').replace(']', '')
        #print('this is clean_annot: ', clean_annot)
        #tokenizing:
        tokens_annot = nltk.word_tokenize(clean_annot)
        pos_tags = nltk.pos_tag(tokens_annot)
        #print('this is pos tags: ', pos_tags)
        tweet_tokens.append(pos_tags)
    #print('this is tweet_tokens: ', tweet_tokens)
    #print('tweet_tokens type: ', type(tweet_tokens))
    #for tag in tweet_tokens:
        #print('in-scope tokens: ', *tag)
        #print('token type: ', type(tag[0]))
    return tweet_tokens

def get_POS_tags(tweet_tokens):
    pos_list = []
    for list_tuples in tweet_tokens:
        for tuples in list_tuples:
            #print(*element[1])
            #extract the second element of each tuple (the POS value):
            pos_tag = [tuples[1] for tuples in list_tuples]
            #print(*pos_tag)
        #append each POS value to the pos_list:
        pos_list.append(pos_tag)
    #sort all the lists of POS values by the number of values (so a list with 2 POS values
    #will show before a list with 3 POS values):
    new_list = sorted(pos_list, key=len)
    print(*new_list, sep='\n')        
    return new_list



if __name__ == '__main__':
    ##I'll use these raw_input functions in the final version of the program    
    # workbook_input = raw_input('enter spreadsheet name (should be xlsx): ') # open files
    # workbook = load_workbook('./' + workbook_input + '.xlsx')
    # sheet_input1 = raw_input('enter annotated spreadsheet name (as it appears on the tab name): ')
    # sheet_input2 = raw_input('enter master list spreadsheet name (as it appears on the tab name): ')
    # sheet1 = workbook[sheet_input1]
    # sheet2 = workbook[sheet_input2]


    workbook = load_workbook('./Scope-Annotation-TrainingDONE.xlsx')

    sheet1 = workbook['ScopeTrainingAnnotNORMA']

    excel_data1 = get_excel_data(sheet1)

    # Extracting tweet_text and proj_id:
    tweet_text = find_tweet_text(excel_data1)

    # Finding scopes:
    scope_found1 = find_scope(tweet_text)

    scope_clean1 = scope_cleaning_tokenizing(scope_found1)

    pos_tags_only = get_POS_tags(scope_clean1)





    #print(scope_clean1)
    # scope_clean_pos_only = []
    # pos_only = [x[1] for x in scope_clean1]
    #print('this is pos_only: ', pos_only)
    #scope_clean_pos_only.append(pos_only)
    #print('this is scope_clean: ', scope_clean_pos_only)
 
    #pos_tags = nltk.pos_tag(scope_clean1)

    # Matching scopes:
    # scope_agreement = scope_match(scope_clean1, scope_clean2)
    # print(scope_agreement)
